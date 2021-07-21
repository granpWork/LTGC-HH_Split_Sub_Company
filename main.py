import os

import openpyxl
import pandas as pd

from datetime import datetime
from pathlib import Path
from openpyxl.styles import Border, Side
from Utils import Utils


def set_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    pass


def getExcelData(fPath, logPath):
    print("Reading: " + fPath + "......")
    print("==============================================================")

    errtmpList = []
    util = Utils()

    df = pd.read_excel(fPath, sheet_name='Eligible Population', header=1,
                       dtype={'PhilHealth_ID*': str, 'Contact_number_of_employer*': str,
                              'Contact_No.*': str, 'Age': str}, na_filter=False)

    totalRecords = len(df.index)
    # groups = df.groupby('Company Name')
    groups = df.groupby('Parent Company')

    for comp, records in groups:
        # outputFileName = comp + "_AZ"
        outputFileName = comp + "_HHLTGC_CEIRMasterlist_PNB_AZ_Switchers"
        records = records.astype(str)

        # get num rows
        numrows = len(records.index)

        errtmpList.append(comp + "  have " + str(numrows) + " records")
        print(comp + "  have " + str(numrows) + " records", end='')

        # print(comp + " --> has " + str(numrows) + " records")

        templateFile = util.duplicateTemplateLTGC(templateFilePath, outPath, outputFileName)

        theFile = openpyxl.load_workbook(templateFile)
        currentSheet = theFile["Eligible Population"]
        util.addingDataValidation(currentSheet, numrows)

        set_border(currentSheet, "A3:BM" + str(numrows + 2))

        theFile.save(templateFile)

        writer = pd.ExcelWriter(templateFile, engine='openpyxl', mode='a')
        writer.book = openpyxl.load_workbook(templateFile)
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
        records.to_excel(writer, sheet_name="Eligible Population", startrow=2, header=False, index=False)
        writer.save()

        print("...Done...")

    errtmpList.append("***")
    errtmpList.append("total Records: " + str(totalRecords))
    errtmpList.append("***")

    print("total Records: " + str(totalRecords))

    util.createLogFile(logPath, errtmpList)

    pass


def getCountResult(outPath):
    arrOutFiles = os.listdir(outPath)

    arrNumRowsCount = []
    for filename in arrOutFiles:
        filenamePath = os.path.join(outPath, filename)
        if not filename == ".DS_Store":
            print(filename)
            df = pd.read_excel(filenamePath)

            # Get Data count per excel file
            arrNumRowsCount.append(len(df.index)-1)

    print(sum(arrNumRowsCount))
    print(arrNumRowsCount)


    pass


def getInFileCout(FilePath):
    df = pd.read_excel(FilePath, sheet_name='Eligible Population', header=1,
                       dtype={'PhilHealth_ID*': str, 'Contact_number_of_employer*': str,
                              'Contact_No.*': str, 'Age': str}, na_filter=False)

    print(len(df.index))

    pass


if __name__ == '__main__':
    pd.set_option('display.max_columns', None)
    pd.set_option('display.max_rows', None)
    pd.set_option('display.width', None)

    today = datetime.today()
    dateTime = today.strftime("%m%d%y%H%M%S")

    dirPath = r"/Users/Ran/Documents/Vaccine/LTGSplit_sub-comp"

    inPath = os.path.join(dirPath, "in")
    outPath = os.path.join(dirPath, "out/hh")
    logPath = os.path.join(dirPath, "log")
    templateFilePath = os.path.join(dirPath, "template/HHLTGC_CEIRMasterlist_ExtraCols.xlsx")

    print("==============================================================")
    print("Running Scpirt: Split Sub Companies HH......")
    print("==============================================================")

    # Get all filenames from folder and convert to list
    arrFilenames = os.listdir(inPath)

    for filename in arrFilenames:
        keyCeirMaster = filename.split("_")[0]

        if not filename == ".DS_Store" and keyCeirMaster.__contains__("HH"):
            FilePath = os.path.join(inPath, filename)

            getExcelData(FilePath, logPath)
            getInFileCout(FilePath)
            getCountResult(outPath)
